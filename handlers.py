import asyncio
import warnings

from aiogram.types import Message, BotCommand
from aiogram.dispatcher.storage import FSMContext
from aiogram.utils.markdown import hbold, hunderline

from main import dp, bot, path, title, flag
from main import TEAM_ID, ADMIN_ID, PROG_ID
from keyboards import *
from states import *
from words import *
from function import *

# pip install pandas openpyxl xlsxwriter xlrd
from openpyxl import load_workbook
from openpyxl import __version__


# Check "openpyxl" version to Errors
if __version__.startswith('3.'):
    warnings.filterwarnings("ignore", message="Conditional Formatting extension is not supported and will be removed",
                            category=UserWarning)


"""                                  ____          _      _____              _ 
                                    | __ )   ___  | |_   | ____|__  __  ___ | |
                                    |  _ \  / _ \ | __|  |  _|  \ \/ / / _ \| |
                                    | |_) || (_) || |_   | |___  >  < |  __/| |
                                    |____/  \___/  \__|  |_____|/_/\_\ \___||_|                                      
"""


# ------------------------------------ Base handlers (start, cancel) ---------------------------------------

@dp.message_handler(commands=["start"], state="*")
async def command_start(message: Message, state: FSMContext) -> None:
    if message.from_user.id not in ADMIN_ID:
        pass
    else:
        await bot.set_my_commands([BotCommand("start", "Запуск/Перезапуск бота"),
                                   BotCommand("download_exel", "Загрузка Exel файла"),
                                   BotCommand("find_product", "Поиск продукта")])
        user = f"@{message.from_user.full_name}" if message.from_user.username is None else f"@{message.from_user.username}"
        await bot.send_photo(chat_id=message.from_user.id, photo=open(f"photos/start_photo.png", "rb"),
                             caption=hbold(f"👋 Привет {user}\n") + START_DESCRIPTION, protect_content=True,
                             reply_markup=get_start_kb())
        await message.delete()
        await state.finish()


@dp.message_handler(text=CANCEL, state="*")
async def command_cancel(message: Message, state: FSMContext) -> None:
    global flag
    if state is None:
        pass
    flag = False
    await state.finish()
    await message.answer(hunderline(CANCEL_WRITING_DATA), reply_markup=get_start_kb())
    await message.delete()


# ---------------------------------------- Menu handlers ----------------------------------------------

#                                                                                                         Download Exel
@dp.message_handler(text=[DOWNLOAD_EXEL, "/download_exel"], state="*")
async def download_exel(message: Message, state: FSMContext) -> None:
    await state.finish()
    await message.answer(text=hbold(SEND_EXEL_FILE), reply_markup=get_cancel_kb())
    await message.delete()
    await DownloadExel.download_exel.set()

    # Хендлер проверки: является ли message документом
    @dp.message_handler(lambda message: not message.document, state=DownloadExel.download_exel)
    async def check_exel_file(message: Message) -> None:
        await message.answer(text=hunderline(IS_NOT_EXEL_FILE))
        await download_exel(message, state)

    @dp.message_handler(content_types=["document"], state=DownloadExel.download_exel)
    async def save_download_exel(message: Message, state: FSMContext) -> None:
        async with state.proxy() as data:
            data["download_exel"] = message.document.file_name
            name_document = data["download_exel"]
        try:
            global path, title
            await message.answer(text=hbold(WAITING_DOWNLOAD))
            await message.document.download(destination_file=f"files/{name_document}")
            path = name_document
            title = load_workbook(f"files/{path}").sheetnames[0]
            for prog_id in PROG_ID:
                await bot.send_document(chat_id=prog_id, document=open(f"files/{path}", "rb"))
            await bot.send_document(chat_id=TEAM_ID, document=open(f"files/{path}", "rb"))
            await message.answer(text=hbold(ENTRY_EXCHANGE_DOLLAR))
            await DownloadExel.next()

        except Exception as error:
            await message.answer(hunderline(ERROR_OPENED_FILE) if str(error).find("Permission denied") >= 0
                                 else hunderline(ERROR_ANOTHER) + str(error), reply_markup=get_start_kb())

    @dp.message_handler(state=DownloadExel.dollar_exchange_rate)
    async def save_dollar_exchange_rate(message: Message, state: FSMContext) -> None:
        async with state.proxy() as data:
            data["dollar_exchange_rate"] = message.text
            exchange_dollar = data["dollar_exchange_rate"]
        global dollar_rate
        dollar_rate = int("".join([str(num) for num in exchange_dollar if num.isdigit()]))
        await message.answer(text=hbold(DATA_DOWNLOADED), reply_markup=get_start_kb())
        await state.finish()


#                                                                                                           Find Values
@dp.message_handler(text=[FIND_VALUES, "/find_product"], state="*")
async def find_values(message: Message, state: FSMContext) -> None:
    global path, title
    if path and title:
        await state.finish()
        await message.answer(text=f"{hbold(ENTRY_SELECTED_PRODUCT)}", reply_markup=get_cancel_kb())
        await FindProduct.find_product.set()
    else:
        await message.answer(text=hunderline(NOT_SELECTED_FILE_EXCEL), reply_markup=get_start_kb())

    @dp.message_handler(state=FindProduct.find_product)
    async def output_find_values(message: Message, state: FSMContext) -> None:
        async with state.proxy() as data:
            data["find_product"] = message.text
            text_find = data["find_product"]

        global path, title, flag, dollar_rate
        listbook = load_workbook(f"files/{path}", data_only=True)[title]
        listbook_profit = load_workbook(f"files/{path}", read_only=True, data_only=True)["data"]

        try:
            out_list: list = []
            products: dict = await get_products(listbook=listbook, text=text_find)  # dict: {row: [a, b, c]}
            categories: list = await get_categories(listbook=listbook)  # list: [a, b, c]

            if not products:
                raise Exception("Message text is empty")

            else:
                for row, list_value in products.items():

                    category = [list_value[num] for num in [2, 9, 5, 8, 4, 6]]
                    providers = {categories[num]: list_value[num] for num in range(10, 22)}

                    profits = await get_profits(listbook=listbook_profit)
                    min_price = await get_minimal_price(providers=providers, profits=profits)

                    if min_price == 0:
                        continue
                    else:
                        category_text: str = ""
                        provider_price = [f"\n\n<b>⭕️ {name}: {p}</b>" for name, p in providers.items() if bool(p)]
                        price_text = "".join(provider_price) + \
                                     f"\n\n<i>🔝 Мин. цена: $ {min_price}</i>" + \
                                     f"\n\n<b>💱 Мин. цена — {min_price * dollar_rate:,d} сум</b>" + \
                                     f"\n<b>💵 Курс доллара — {dollar_rate:,d} сум</b>"

                        for text in category:
                            if text is None or text == 0:
                                continue
                            category_text += f"\n<b>{text}</b>"
                        out_list.append(category_text + price_text)
                else:
                    if len(out_list) == 0:
                        raise Exception("Message text is empty")

            flag = True
            for text in out_list:
                if not flag:
                    break
                await message.answer(text=text, reply_markup=get_cancel_kb())
                await asyncio.sleep(1.5)

        except Exception as error:
            if str(error) == "Message text is empty":
                await message.answer(text=hunderline(ERROR_MESSAGE_EMPTY))  # Error №1: Message text is empty
            elif str(error).find("Flood control exceeded") >= 0:
                await message.answer(text=hunderline(ERROR_FLOOD))  # Error №2: Flood control exceeded.
            else:
                await message.answer(text=hunderline(ERROR_ANOTHER) + str(error))  # Error №3: Another errors
        finally:
            await state.finish()
            await find_values(message, state)
